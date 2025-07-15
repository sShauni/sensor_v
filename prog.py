import tkinter as tk
from tkinter import messagebox
from datetime import datetime
from openpyxl import Workbook, load_workbook
import threading
import os
import time
import platform

# === VARIÁVEIS GLOBAIS ===
# PRIMEIRO definimos os pinos dos sensores
SENSOR_LARGADA = 26 #antes 17
SENSOR_CHEGADA = 20 #antes 27

# Depois as configurações que dependem dos pinos
# === CONFIGURAÇÕES ===
if platform.system() == 'Windows':
    TESTE_TOQUE = True  # Ativa simulação por toque no Windows
    USANDO_GPIO = False
    USANDO_TFT = False
else:
    TESTE_TOQUE = False
    USANDO_GPIO = True
    USANDO_TFT = True
        
if USANDO_GPIO:
    try:
        import RPi.GPIO as GPIO
        GPIO.setmode(GPIO.BCM)
        # Configuração especial para evitar conflitos com TFT
        if USANDO_TFT:
            GPIO.setup(SENSOR_LARGADA, GPIO.IN, pull_up_down=GPIO.PUD_UP)  # Invertida a lógica
            GPIO.setup(SENSOR_CHEGADA, GPIO.IN, pull_up_down=GPIO.PUD_UP)  # Invertida a lógica
        else:
            GPIO.setup(SENSOR_LARGADA, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)
            GPIO.setup(SENSOR_CHEGADA, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)
    except Exception as e:
        messagebox.showerror("Erro de GPIO", f"Erro na configuração: {str(e)}")
        USANDO_GPIO = False

# Restante das variáveis globais
sensor_simulado_largada = False
sensor_simulado_chegada = False
PASTA_LOGS = 'logs'
log_index = 0
ARQUIVO_EXCEL = ''
lista_logs = []
log_atual = ''
leituras = []
medindo = False
thread_leitura = None

print(f"TESTE_TOQUE: {TESTE_TOQUE}, USANDO_GPIO: {USANDO_GPIO}")
# ===================================================================
#  INÍCIO DO BLOCO DE TODAS AS FUNÇÕES
#  (Movido para antes da criação da GUI para evitar NameError)
# ===================================================================

def listar_logs_existentes():
    global lista_logs, btn_ant, btn_prox
    if not os.path.exists(PASTA_LOGS):
        os.makedirs(PASTA_LOGS)
    lista_logs = sorted([f[:-5] for f in os.listdir(PASTA_LOGS) if f.endswith('.xlsx')])
    # A atualização dos botões é chamada dentro de carregar_log_existente ou criar_novo_log
    # para garantir que os botões já existam.

def criar_novo_log():
    global ARQUIVO_EXCEL, leituras, log_atual, log_index
    hoje = datetime.now().strftime('%d%m%y')
    listar_logs_existentes() # Atualiza a lista de logs primeiro
    existentes = [f for f in lista_logs if f.startswith(hoje)]
    novo_nome = hoje if not existentes else f"{hoje}-{len(existentes) + 1}"
    
    log_atual = novo_nome
    ARQUIVO_EXCEL = os.path.join(PASTA_LOGS, f"{log_atual}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(['Data', 'Hora', 'Passagem', 'Tempo (s)'])
    wb.save(ARQUIVO_EXCEL)
    
    listar_logs_existentes() # Re-lista para incluir o novo arquivo
    log_index = lista_logs.index(log_atual)
    
    leituras = []
    atualizar_lista()
    atualizar_titulo_log()
    atualizar_botoes_navegacao()

def carregar_log_existente(index):
    global ARQUIVO_EXCEL, log_index, log_atual, leituras
    if 0 <= index < len(lista_logs):
        log_index = index
        log_atual = lista_logs[log_index]
        ARQUIVO_EXCEL = os.path.join(PASTA_LOGS, f"{log_atual}.xlsx")
        try:
            wb = load_workbook(ARQUIVO_EXCEL)
            ws = wb.active
            leituras = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] is not None:
                    leituras.append({'data': row[0], 'hora': row[1], 'tempo': row[3], 'passagem': row[2]})
            
            atualizar_titulo_log()
            atualizar_lista()
        except Exception as e:
            messagebox.showerror("Erro ao carregar Log", f"Não foi possível carregar o arquivo {log_atual}.xlsx: {e}")
            leituras = []
    
    atualizar_botoes_navegacao()

def atualizar_botoes_navegacao():
    btn_ant.config(state=tk.NORMAL if log_index > 0 else tk.DISABLED)
    btn_prox.config(state=tk.NORMAL if log_index < len(lista_logs) - 1 else tk.DISABLED)

def atualizar_titulo_log():
    lbl_log_titulo.config(text=log_atual)

def salvar_leitura_excel(leitura):
    print(f"[SALVAR] {leitura}")

    global ARQUIVO_EXCEL
    
    try:
        # Verifica se o arquivo já existe
        if os.path.exists(ARQUIVO_EXCEL):
            wb = load_workbook(ARQUIVO_EXCEL)
            ws = wb.active
        else:
            # Se não existir, cria um novo
            wb = Workbook()
            ws = wb.active
            ws.append(['Data', 'Hora', 'Passagem', 'Tempo (s)'])  # Cabeçalhos
        
        # Adiciona os dados da leitura
        ws.append([
            leitura['data'],
            leitura['hora'],
            leitura['passagem'],
            leitura['tempo']
        ])
        
        # Salva o arquivo
        wb.save(ARQUIVO_EXCEL)
        print(f"Leitura {leitura['passagem']} salva com sucesso em {ARQUIVO_EXCEL}")
        
    except Exception as e:
        print(f"ERRO ao salvar leitura: {str(e)}")
        messagebox.showerror("Erro ao Salvar", f"Não foi possível salvar a leitura: {str(e)}")

def leitura_id():
    existente = [l['passagem'] for l in leituras]
    return max(existente + [0]) + 1
    
def atualizar_circulo(sensor, ativo):
    if sensor == SENSOR_LARGADA:
        canvas_largada.itemconfig(circ_largada, fill='red' if ativo else 'darkgreen')
    elif sensor == SENSOR_CHEGADA:
        canvas_chegada.itemconfig(circ_chegada, fill='red' if ativo else 'darkgreen')
    root.update()  # Força atualização da interface    

def simular_sensor(sensor_id):
    global sensor_simulado_largada, sensor_simulado_chegada
    print(f"[SIMULAÇÃO] Sensor {sensor_id} clicado")

    if not TESTE_TOQUE or not medindo:
        print("[SIMULAÇÃO] Ignorado - não está medindo ou não em teste")
        return

    if sensor_id == SENSOR_LARGADA:
        sensor_simulado_largada = True
        atualizar_circulo(SENSOR_LARGADA, True)
        root.after(100, lambda: atualizar_circulo(SENSOR_LARGADA, False))  # só visual
        root.after(100, lambda: desativar_sensor_simulado('largada'))      # desativa variável

    elif sensor_id == SENSOR_CHEGADA:
        sensor_simulado_chegada = True
        atualizar_circulo(SENSOR_CHEGADA, True)
        root.after(100, lambda: atualizar_circulo(SENSOR_CHEGADA, False))  # só visual
        root.after(100, lambda: desativar_sensor_simulado('chegada'))      # desativa variável

def desativar_sensor_simulado(sensor):
    global sensor_simulado_largada, sensor_simulado_chegada
    if sensor == 'largada':
        sensor_simulado_largada = False
    elif sensor == 'chegada':
        sensor_simulado_chegada = False

def registrar_tempo(t1, t2):
    print(">>> Registrando tempo")
    tempo = round(t2 - t1, 2)
    leitura = {
        'data': datetime.now().strftime('%d/%m/%Y'),
        'hora': datetime.now().strftime('%H:%M:%S'),
        'tempo': tempo,
        'passagem': leitura_id()
    }
    leituras.append(leitura)
    salvar_leitura_excel(leitura)
    
    # Atualiza a interface na thread principal
    root.after(0, atualizar_lista)
    
    # Pisca o LED de chegada
    for _ in range(3):
        if not medindo: break
        root.after(0, lambda: atualizar_circulo(SENSOR_CHEGADA, True))
        time.sleep(0.2)
        root.after(0, lambda: atualizar_circulo(SENSOR_CHEGADA, False))
        time.sleep(0.2)

def tratar_largada():
    global sensor_simulado_largada, sensor_simulado_chegada
    print("[LARGADA] Tratando largada...")
    
    # Espera largada ser ativada
    while medindo:
        ativo = (not GPIO.input(SENSOR_LARGADA)) if USANDO_TFT else GPIO.input(SENSOR_LARGADA)
        if TESTE_TOQUE and sensor_simulado_largada:
            ativo = True
        if ativo:
            break
        time.sleep(0.01)
        root.update_idletasks()
    
    t1 = time.time()  # Começa a contar exatamente na ativação da largada
    atualizar_circulo(SENSOR_LARGADA, True)

    # Espera largada ser liberada para evitar múltiplas leituras
    while medindo:
        ativo = GPIO.input(SENSOR_LARGADA) if USANDO_TFT else not GPIO.input(SENSOR_LARGADA)
        if TESTE_TOQUE and not sensor_simulado_largada:
            ativo = False
        if not ativo:
            break
        time.sleep(0.01)
        root.update_idletasks()

    atualizar_circulo(SENSOR_LARGADA, False)
    time.sleep(0.1)

    # Aguarda chegada (até 10 segundos)
    tempo_inicio = time.time()
    while medindo and (time.time() - tempo_inicio < 10):
        ativo = GPIO.input(SENSOR_CHEGADA) if USANDO_TFT else not GPIO.input(SENSOR_LARGADA)
        if TESTE_TOQUE and sensor_simulado_chegada:
            ativo = True
        if ativo:
            registrar_tempo(t1, time.time())
            if TESTE_TOQUE:
                sensor_simulado_chegada = False
            break
        time.sleep(0.01)
        root.update_idletasks()

    atualizar_circulo(SENSOR_CHEGADA, False)
    if TESTE_TOQUE:
        sensor_simulado_largada = False
        sensor_simulado_chegada = False

def monitorar():
    global medindo
    print("[MONITORAR] Iniciando monitoramento...")    
    try:
        while medindo:
            # Manter a interface responsiva
            root.update_idletasks()
            
            if TESTE_TOQUE:
                # Lógica de simulação
                if sensor_simulado_largada:
                    tratar_largada()
                    while medindo and sensor_simulado_largada:
                        time.sleep(0.01)
                        root.update_idletasks()
            elif USANDO_GPIO:
                # Lógica com GPIO real
                estado_largada = GPIO.input(SENSOR_LARGADA) if USANDO_TFT else not GPIO.input(SENSOR_LARGADA)
                if estado_largada:
                    tratar_largada()
            
            time.sleep(0.05)  # Pequena pausa para não sobrecarregar
    except Exception as e:
        print(f"Erro: {e}")
    finally:
        medindo = False
        btn_iniciar.after(100, lambda: btn_iniciar.config(
            text="INICIAR", 
            bg='lightgreen', 
            fg='black', 
            state=tk.NORMAL
        ))

def atualizar_lista():
    lista.delete(0, tk.END)
    for l in sorted(leituras, key=lambda x: x['passagem']):
        lista.insert(tk.END, f"{l['passagem']:02} - {l['tempo']}s")

def iniciar_parar():
    global medindo, thread_leitura
    print(f"[BOTÃO] medindo: {medindo}")    
    if not medindo:
        # Iniciar monitoramento
        if not ARQUIVO_EXCEL:
            messagebox.showwarning("Sem Log", "Crie um novo log antes de iniciar a cronometragem.")
            return
            
        medindo = True
        btn_iniciar.config(text="PARAR", bg='red', fg='white', activebackground='darkred')
        print("[BOTÃO] Iniciando thread...")
        thread_leitura = threading.Thread(target=monitorar)
        thread_leitura.daemon = True
        thread_leitura.start()
    else:
        # Parar monitoramento
        medindo = False
        btn_iniciar.config(text="INICIAR", bg='lightgreen', fg='black', state=tk.NORMAL)

def excluir_leitura():
    sel = lista.curselection()
    if not sel:
        messagebox.showinfo("Nenhuma Seleção", "Por favor, selecione uma leitura para excluir.")
        return
    
    passagem_a_remover_str = lista.get(sel[0]).split(' ')[0]
    passagem_a_remover = int(passagem_a_remover_str)
    
    leitura_idx_real = next((i for i, l in enumerate(leituras) if l['passagem'] == passagem_a_remover), -1)
            
    if leitura_idx_real == -1:
        messagebox.showerror("Erro", "Não foi possível encontrar a leitura para excluir.")
        return

    if messagebox.askyesno("Confirmar Exclusão", f"Tem certeza que deseja excluir a passagem {passagem_a_remover:02}?"):
        try:
            wb = load_workbook(ARQUIVO_EXCEL)
            ws = wb.active
            row_to_delete = -1
            for i in range(ws.max_row, 1, -1):
                if ws.cell(row=i, column=3).value == passagem_a_remover:
                    row_to_delete = i
                    break
            
            if row_to_delete != -1:
                ws.delete_rows(row_to_delete)
                wb.save(ARQUIVO_EXCEL)
                leituras.pop(leitura_idx_real)
                atualizar_lista()
            else:
                messagebox.showwarning("Erro de Exclusão", "Leitura não encontrada na planilha Excel.")

        except Exception as e:
            messagebox.showerror("Erro ao Excluir", f"Erro ao excluir leitura do Excel: {e}")

def inicializar_app():
    listar_logs_existentes()
    if lista_logs:
        carregar_log_existente(len(lista_logs) - 1)
    else:
        criar_novo_log()
# ===================================================================
#  FIM DO BLOCO DE FUNÇÕES
# ===================================================================


# === CONFIGURAÇÃO GPIO ===
if USANDO_GPIO:
    try:
        GPIO.setmode(GPIO.BCM)
        GPIO.setup(SENSOR_LARGADA, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)
        GPIO.setup(SENSOR_CHEGADA, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)
    except Exception as e:
        messagebox.showerror("Erro de GPIO", f"Não foi possível configurar os pinos GPIO: {e}")
        USANDO_GPIO = False

# === GUI (Interface Gráfica) ===
root = tk.Tk()
root.title("Cronometragem")
root.attributes('-fullscreen', True)
root.configure(bg='white')

frame_lista = tk.Frame(root, bg='white')
frame_lista.place(x=0, y=0, width=240, height=320)

frame_top = tk.Frame(frame_lista, bg='white')
frame_top.pack()

btn_ant = tk.Button(frame_top, text="<", command=lambda: carregar_log_existente(log_index - 1), font=("Arial", 10))
btn_ant.pack(side=tk.LEFT)

lbl_log_titulo = tk.Label(frame_top, text="", font=("Arial", 12, 'bold'), bg='white')
lbl_log_titulo.pack(side=tk.LEFT, padx=5)

btn_prox = tk.Button(frame_top, text=">", command=lambda: carregar_log_existente(log_index + 1), font=("Arial", 10))
btn_prox.pack(side=tk.LEFT)

lista = tk.Listbox(frame_lista, font=("Arial", 14), selectmode=tk.SINGLE)
lista.pack(fill=tk.BOTH, expand=True)

frame_controles = tk.Frame(root, bg='white')
frame_controles.place(x=240, y=0, width=240, height=320)

btn_iniciar = tk.Button(
    frame_controles, 
    text="INICIAR", 
    bg='lightgreen', 
    fg='black', 
    font=("Arial", 14), 
    width=10, 
    command=iniciar_parar,
    activebackground='red',  # Cor quando pressionado
    activeforeground='white'  # Texto quando pressionado
)
btn_iniciar.pack(pady=5)

btn_novo_log = tk.Button(frame_controles, text="NOVO LOG", bg='blue', fg='white', font=("Arial", 10), command=criar_novo_log)
btn_novo_log.pack(pady=5)

btn_excluir = tk.Button(frame_controles, text="EXCLUIR", bg='red', fg='white', font=("Arial", 10), command=excluir_leitura)
btn_excluir.pack(pady=5)

lbl_largada = tk.Label(frame_controles, text="LARGADA", font=("Arial", 10), bg='white')
lbl_largada.pack()
canvas_largada = tk.Canvas(frame_controles, width=50, height=50, bg='white', highlightthickness=0)
circ_largada = canvas_largada.create_oval(5, 5, 45, 45, fill='darkgreen')
if TESTE_TOQUE:
    canvas_largada.bind("<Button-1>", lambda e: simular_sensor(SENSOR_LARGADA))
canvas_largada.pack()

lbl_chegada = tk.Label(frame_controles, text="CHEGADA", font=("Arial", 10), bg='white')
lbl_chegada.pack()
canvas_chegada = tk.Canvas(frame_controles, width=50, height=50, bg='white', highlightthickness=0)
circ_chegada = canvas_chegada.create_oval(5, 5, 45, 45, fill='darkgreen')
if TESTE_TOQUE:
    canvas_chegada.bind("<Button-1>", lambda e: simular_sensor(SENSOR_CHEGADA))
canvas_chegada.pack()


# === INICIALIZAÇÃO E LOOP PRINCIPAL ===
inicializar_app()
root.mainloop()

# === LIMPEZA AO FECHAR ===
if USANDO_GPIO:
    GPIO.cleanup()